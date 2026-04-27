from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, make_response
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from models import db, Product, Appointment, Invoice, InvoiceItem, User, Client, CashRegister, InventoryMovement, Company, ProductComponent, ProductBatch
from datetime import datetime, timedelta, timezone
from functools import wraps

def get_now():
    # Fuerza la hora de Bogotá (UTC-5) independientemente del servidor
    return datetime.now(timezone(timedelta(hours=-5))).replace(tzinfo=None)
import os
import io
import csv
import zipfile
from flask import send_file, make_response
from openpyxl import Workbook
from openpyxl.styles import Font

app = Flask(__name__)
# Configuración básica
app.config['SECRET_KEY'] = 'dev-secret-key-change-in-production'
basedir = os.path.abspath(os.path.dirname(__name__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'erp_system.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(basedir, 'static', 'uploads')

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Inicializar SQLAlchemy con la app
db.init_app(app)

# Configurar Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor inicia sesión para acceder a esta página.'
login_manager.login_message_category = 'warning'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@app.template_filter('currency')
def currency_filter(value):
    if value is None:
        return "$0"
    try:
        # Si el valor es entero o muy cercano a entero, no mostrar decimales
        if abs(value - round(value)) < 0.01:
            return f"${int(round(value)):,}".replace(",", ".")
        else:
            # Mostrar 2 decimales si hay centavos significativos
            formatted = f"{value:,.2f}"
            return f"${formatted.replace(',', 'X').replace('.', ',').replace('X', '.')}"
    except (ValueError, TypeError):
        return f"${value}"

@app.context_processor
def inject_company():
    if current_user.is_authenticated:
        return dict(company=current_user.company)
    return dict(company=None)

def permission_required(permission_name):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if current_user.role == 'admin':
                return f(*args, **kwargs)
            if not current_user.permissions or permission_name not in current_user.permissions:
                flash('No tienes permisos para acceder a este módulo.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@app.before_request
def check_security_and_subscription():
    if current_user.is_authenticated:
        # Check God Mode routes - skip normal checks
        if request.endpoint == 'system_admin':
            return
            
        company = current_user.company
        if company:
            # Auto-suspend if expired
            if get_now() > company.valid_until and company.subscription_status != 'suspended':
                company.subscription_status = 'suspended'
                db.session.commit()
                
            # Restrict access if suspended
            if company.subscription_status == 'suspended':
                allowed_endpoints = ['suspended', 'export_data', 'logout', 'static', 'system_admin', 'login']
                if request.endpoint not in allowed_endpoints:
                    return redirect(url_for('suspended'))
            
            # Setup Wizard check (New companies need to be set up)
            if not company.document_id and request.endpoint not in ['setup_company', 'logout', 'static']:
                return redirect(url_for('setup_company'))
                    
        # Force password change
        if current_user.must_change_password and request.endpoint not in ['change_password', 'logout', 'static', 'suspended', 'setup_company']:
            flash('Por seguridad, debes cambiar tu contraseña inicial antes de continuar.', 'warning')
            return redirect(url_for('change_password'))

# ==========================================
# RUTAS PÚBLICAS Y AUTENTICACIÓN
# ==========================================
@app.route('/')
def landing():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('landing.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        email = request.form.get('username') # HTML input name is still username
        password = request.form.get('password')
        user = User.query.filter_by(email=email).first()
        
        if user and check_password_hash(user.password_hash, password):
            if not user.is_active:
                flash('Tu cuenta ha sido suspendida. Contacta al administrador.', 'danger')
                return redirect(url_for('login'))
                
            login_user(user)
            if user.must_change_password:
                return redirect(url_for('change_password'))
                
            next_page = request.args.get('next')
            return redirect(next_page if next_page else url_for('dashboard'))
        else:
            flash('Correo electrónico o contraseña incorrectos', 'danger')
            
    return render_template('login.html')

@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    if not current_user.must_change_password:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        if new_password != confirm_password:
            flash('Las contraseñas no coinciden.', 'danger')
        elif len(new_password) < 6:
            flash('La contraseña debe tener al menos 6 caracteres.', 'danger')
        else:
            current_user.password_hash = generate_password_hash(new_password)
            current_user.must_change_password = False
            db.session.commit()
            flash('Contraseña actualizada correctamente. ¡Bienvenido!', 'success')
            return redirect(url_for('dashboard'))
            
    return render_template('change_password.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('landing'))



@app.route('/admin/users')
@login_required
def admin_users():
    if current_user.role != 'admin':
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('dashboard'))
    users = User.query.filter_by(company_id=current_user.company_id).all()
    return render_template('users.html', users=users)

@app.route('/admin/users/add', methods=['POST'])
@login_required
def add_user():
    if current_user.role != 'admin':
        return redirect(url_for('dashboard'))
        
    email = request.form.get('email')
    name = request.form.get('name')
    password = request.form.get('password')
    permissions = ','.join(request.form.getlist('permissions'))
    
    if User.query.filter_by(email=email).first():
        flash('El correo ya está registrado.', 'danger')
        return redirect(url_for('admin_users'))
        
    new_user = User(
        email=email,
        name=name,
        password_hash=generate_password_hash(password),
        role='employee',
        permissions=permissions,
        is_active=True,
        must_change_password=True,
        company_id=current_user.company_id
    )
    db.session.add(new_user)
    db.session.commit()
    flash(f'Usuario {name} creado. Su contraseña inicial es {password}', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/toggle/<int:id>', methods=['POST'])
@login_required
def toggle_user(id):
    if current_user.role != 'admin':
        return redirect(url_for('dashboard'))
    user = User.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    if user.id == current_user.id:
        flash('No puedes inhabilitarte a ti mismo.', 'danger')
    else:
        user.is_active = not user.is_active
        db.session.commit()
        status = 'habilitado' if user.is_active else 'inhabilitado'
        flash(f'Usuario {user.name} ha sido {status}.', 'success')
    return redirect(url_for('admin_users'))

# ==========================================
# DASHBOARD
# ==========================================
@app.route('/dashboard')
@login_required
def dashboard():
    total_products = Product.query.filter_by(company_id=current_user.company_id).count()
    low_stock_list = Product.query.filter(
        Product.company_id == current_user.company_id,
        Product.track_stock == True,
        Product.stock <= Product.min_stock
    ).all()
    low_stock_products = len(low_stock_list)
    
    today = get_now().date()
    appointments_today = Appointment.query.filter(
        Appointment.company_id == current_user.company_id,
        db.func.date(Appointment.date) == today
    ).count()
    total_appointments = Appointment.query.filter_by(company_id=current_user.company_id).count()

    # Expiring products
    expiring_soon_count = 0
    all_expirable = Product.query.filter(
        Product.company_id == current_user.company_id,
        Product.expiry_date.isnot(None)
    ).all()
    for p in all_expirable:
        if (p.expiry_date - today).days <= p.expiry_alert_days:
            expiring_soon_count += 1

    # Chart Data
    seven_days_ago = today - timedelta(days=7)
    sales_trend = db.session.query(
        db.func.strftime('%d-%m', Invoice.date).label('day'),
        db.func.sum(Invoice.total).label('total')
    ).filter(
        Invoice.company_id == current_user.company_id,
        Invoice.date >= seven_days_ago
    ).group_by('day').order_by(Invoice.date.asc()).all()

    chart_labels = [s.day for s in sales_trend]
    chart_values = [float(s.total) for s in sales_trend]

    top_products = Product.query.filter_by(company_id=current_user.company_id).order_by(Product.stock.asc()).limit(5).all()
    top_labels = [p.name for p in top_products]
    top_values = [p.stock for p in top_products]

    # Datos para Gráfico: Top Clientes (por total de compras)
    top_clients_query = db.session.query(
        Client.name,
        db.func.sum(Invoice.total).label('total_spent')
    ).join(Invoice, Client.id == Invoice.client_id).filter(
        Invoice.company_id == current_user.company_id
    ).group_by(Client.id).order_by(db.desc('total_spent')).limit(5).all()

    client_labels = [c.name for c in top_clients_query]
    client_values = [float(c.total_spent) for c in top_clients_query]

    return render_template('index.html', 
                         total_products=total_products, 
                         low_stock_products=low_stock_products,
                         total_appointments=total_appointments,
                         appointments_today=appointments_today,
                         expiring_soon_count=expiring_soon_count,
                         chart_labels=chart_labels,
                         chart_values=chart_values,
                         top_labels=top_labels,
                         top_values=top_values,
                         client_labels=client_labels,
                         client_values=client_values)

# ==========================================
# MÓDULO: INVENTARIO
# ==========================================
@app.route('/inventory')
@login_required
@permission_required('inventory')
def inventory():
    page = request.args.get('page', 1, type=int)
    sort = request.args.get('sort', 'id_desc')
    alert_only = request.args.get('alert') == 'true'
    
    query = Product.query.filter_by(company_id=current_user.company_id)
    
    if alert_only:
        query = query.filter(Product.track_stock == True, Product.stock <= Product.min_stock)
    
    if sort == 'stock_asc':
        query = query.order_by(Product.stock.asc())
    elif sort == 'stock_desc':
        query = query.order_by(Product.stock.desc())
    elif sort == 'price_asc':
        query = query.order_by(Product.sale_price.asc())
    elif sort == 'price_desc':
        query = query.order_by(Product.sale_price.desc())
    elif sort == 'expiry_soon':
        query = query.order_by(Product.expiry_date.asc())
    else:
        query = query.order_by(Product.id.desc())
        
    pagination = query.paginate(page=page, per_page=10, error_out=False)
    products = pagination.items
    return render_template('inventory.html', products=products, pagination=pagination, today=get_now().date())

@app.route('/inventory/add', methods=['POST'])
@login_required
@permission_required('inventory')
def add_product():
    name = request.form.get('name')
    description = request.form.get('description')
    purchase_price = float(request.form.get('purchase_price', 0))
    sale_price = float(request.form.get('sale_price', 0))
    stock = float(request.form.get('stock', 0))
    min_stock = float(request.form.get('min_stock', 5.0))
    track_stock = 'track_stock' in request.form
    
    expiry_date_str = request.form.get('expiry_date')
    expiry_date = None
    if expiry_date_str:
        try:
            expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d').date()
        except ValueError:
            expiry_date = None
            
    expiry_alert_days = request.form.get('expiry_alert_days')
    expiry_alert_days = int(expiry_alert_days) if expiry_alert_days else 30

    new_product = Product(
        name=name,
        description=description,
        purchase_price=purchase_price,
        sale_price=sale_price,
        stock=stock,
        min_stock=min_stock,
        track_stock=track_stock,
        expiry_date=expiry_date,
        expiry_alert_days=expiry_alert_days,
        company_id=current_user.company_id
    )
    db.session.add(new_product)
    db.session.flush()
    
    if stock > 0:
        movement = InventoryMovement(
            product_id=new_product.id,
            movement_type='in',
            quantity=stock,
            reason='Inventario inicial',
            user_id=current_user.id,
            company_id=current_user.company_id
        )
        db.session.add(movement)
        
    db.session.commit()
    flash('Producto agregado correctamente.', 'success')
    return redirect(url_for('inventory'))

@app.route('/product/<int:id>/recipe')
@login_required
@permission_required('inventory')
def product_recipe(id):
    product = Product.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    all_products = Product.query.filter_by(company_id=current_user.company_id).all()
    return render_template('product_recipe.html', product=product, all_products=all_products)

@app.route('/product/<int:id>/component/add', methods=['POST'])
@login_required
@permission_required('inventory')
def add_product_component(id):
    component_id = request.form.get('component_id')
    quantity = float(request.form.get('quantity', 1.0))
    
    # Validar que no se agregue a sí mismo (ya está en el HTML pero por seguridad)
    if int(component_id) == id:
        flash('Un producto no puede ser insumo de sí mismo.', 'danger')
        return redirect(url_for('product_recipe', id=id))

    new_comp = ProductComponent(
        parent_product_id=id,
        component_product_id=component_id,
        quantity=quantity
    )
    db.session.add(new_comp)
    db.session.commit()
    flash('Insumo agregado correctamente.', 'success')
    return redirect(url_for('product_recipe', id=id))

@app.route('/component/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('inventory')
def delete_product_component(id):
    comp = ProductComponent.query.get_or_404(id)
    # Verificar que el producto padre sea de la empresa del usuario
    parent = Product.query.get(comp.parent_product_id)
    if parent.company_id != current_user.company_id:
        abort(403)
        
    db.session.delete(comp)
    db.session.commit()
    flash('Insumo eliminado.', 'info')
    return redirect(url_for('product_recipe', id=comp.parent_product_id))

@app.route('/inventory/edit/<int:id>', methods=['POST'])
@login_required
@permission_required('inventory')
def edit_product(id):
    product = Product.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    
    new_stock = float(request.form.get('stock', 0))
    if new_stock != product.stock:
        diff = new_stock - product.stock
        m_type = 'in' if diff > 0 else 'out'
        movement = InventoryMovement(
            product_id=product.id,
            movement_type=m_type,
            quantity=abs(diff),
            reason='Ajuste en edición',
            user_id=current_user.id,
            company_id=current_user.company_id
        )
        db.session.add(movement)

    product.name = request.form.get('name')
    product.description = request.form.get('description')
    product.purchase_price = float(request.form.get('purchase_price', 0))
    product.sale_price = float(request.form.get('sale_price', 0))
    product.stock = new_stock
    product.min_stock = float(request.form.get('min_stock', 5.0))
    product.track_stock = 'track_stock' in request.form
    
    expiry_date_str = request.form.get('expiry_date')
    if expiry_date_str:
        try:
            product.expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass # Keep old or set None? User intent is probably to keep if invalid
    else:
        product.expiry_date = None
        
    expiry_alert_days = request.form.get('expiry_alert_days')
    product.expiry_alert_days = int(expiry_alert_days) if expiry_alert_days else 30
    
    db.session.commit()
    flash('Producto actualizado correctamente.', 'success')
    return redirect(url_for('inventory'))

@app.route('/product/<int:id>/batches', methods=['GET', 'POST'])
@login_required
@permission_required('inventory')
def product_batches(id):
    product = Product.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            batch_number = request.form.get('batch_number')
            qty = float(request.form.get('quantity', 0))
            expiry = request.form.get('expiry_date')
            
            if not expiry:
                flash('La fecha de vencimiento es obligatoria para un lote.', 'danger')
            else:
                expiry_date = datetime.strptime(expiry, '%Y-%m-%d').date()
                new_batch = ProductBatch(
                    product_id=id,
                    batch_number=batch_number,
                    quantity=qty,
                    expiry_date=expiry_date,
                    company_id=current_user.company_id
                )
                db.session.add(new_batch)
                
                # Actualizar stock total del producto
                product.stock += qty
                
                # Registrar movimiento
                movement = InventoryMovement(
                    product_id=id,
                    movement_type='in',
                    quantity=qty,
                    reason=f'Nuevo lote {batch_number}',
                    user_id=current_user.id,
                    company_id=current_user.company_id
                )
                db.session.add(movement)
                db.session.commit()
                flash('Lote agregado correctamente.', 'success')
        
        elif action == 'delete':
            batch_id = request.form.get('batch_id')
            batch = ProductBatch.query.filter_by(id=batch_id, company_id=current_user.company_id).first()
            if batch:
                product.stock -= batch.quantity
                movement = InventoryMovement(
                    product_id=id,
                    movement_type='out',
                    quantity=batch.quantity,
                    reason=f'Lote {batch.batch_number} eliminado',
                    user_id=current_user.id,
                    company_id=current_user.company_id
                )
                db.session.add(movement)
                db.session.delete(batch)
                db.session.commit()
                flash('Lote eliminado.', 'info')
                
        return redirect(url_for('product_batches', id=id))
        
    batches = ProductBatch.query.filter_by(product_id=id).order_by(ProductBatch.expiry_date.asc()).all()
    return render_template('product_batches.html', product=product, batches=batches, today=get_now().date())

@app.route('/inventory/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('inventory')
def delete_product(id):
    product = Product.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    db.session.delete(product)
    db.session.commit()
    flash('Producto eliminado correctamente.', 'success')
    return redirect(url_for('inventory'))

@app.route('/inventory/adjust/<int:id>', methods=['POST'])
@login_required
@permission_required('inventory')
def adjust_product(id):
    product = Product.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    diff = int(request.form.get('quantity', 0))
    reason = request.form.get('reason', 'Ajuste manual')
    
    if diff == 0:
        flash('La cantidad a ajustar no puede ser cero.', 'warning')
        return redirect(url_for('inventory'))
        
    m_type = 'in' if diff > 0 else 'out'
    
    if m_type == 'out' and product.stock < abs(diff):
        flash('No hay suficiente stock para descontar esa cantidad.', 'danger')
        return redirect(url_for('inventory'))
        
    product.stock += diff
    movement = InventoryMovement(
        product_id=product.id,
        movement_type=m_type,
        quantity=abs(diff),
        reason=reason,
        user_id=current_user.id,
        company_id=current_user.company_id
    )
    db.session.add(movement)
    db.session.commit()
    flash('Stock ajustado correctamente.', 'success')
    return redirect(url_for('inventory'))

@app.route('/inventory/history')
@login_required
@permission_required('inventory')
def inventory_history():
    query = InventoryMovement.query.filter(InventoryMovement.company_id == current_user.company_id).join(Product)
    filter_type = request.args.get('filter', 'all')
    search = request.args.get('search', '').strip()
    
    if filter_type == 'sales':
        query = query.filter(InventoryMovement.reason.like('Venta Factura%'))
    elif filter_type == 'adjustments':
        query = query.filter(~InventoryMovement.reason.like('Venta Factura%'))
        
    if search:
        if search.isdigit():
            query = query.filter((Product.name.ilike(f'%{search}%')) | (Product.id == int(search)))
        else:
            query = query.filter(Product.name.ilike(f'%{search}%'))
            
    movements = query.order_by(InventoryMovement.date.desc()).all()
    return render_template('inventory_history.html', movements=movements, current_filter=filter_type, current_search=search)

@app.route('/api/products')
@login_required
def api_products():
    products = Product.query.filter_by(company_id=current_user.company_id).all()
    return jsonify([p.to_dict() for p in products])

@app.route('/api/notifications')
@login_required
def get_notifications():
    notifications = []
    
    # 1. Low stock
    if current_user.role == 'admin' or (current_user.permissions and 'inventory' in current_user.permissions):
        low_stock_items = Product.query.filter(
            Product.company_id == current_user.company_id, 
            Product.track_stock == True,
            Product.stock <= Product.min_stock
        ).all()
        for item in low_stock_items:
            notifications.append({
                'type': 'warning',
                'title': 'Stock Bajo',
                'message': f'{item.name} tiene solo {item.stock} unidades.',
                'link': url_for('inventory', alert='true')
            })

    # 2. Expiration alerts
    if current_user.role == 'admin' or (current_user.permissions and 'inventory' in current_user.permissions):
        today = get_now().date()
        expiring_products = Product.query.filter(
            Product.company_id == current_user.company_id,
            Product.expiry_date.isnot(None)
        ).all()
        
        for p in expiring_products:
            days_left = (p.expiry_date - today).days
            if days_left <= p.expiry_alert_days:
                label = "VENCIDO" if days_left < 0 else f"Vence en {days_left} días"
                notifications.append({
                    'type': 'warning',
                    'title': 'Alerta de Vencimiento',
                    'message': f'{p.name} {label} ({p.expiry_date.strftime("%d-%m-%Y")}).',
                    'link': url_for('inventory', sort='expiry_soon') # We will add this sort
                })
            
    # 3. Today's appointments
    if current_user.role == 'admin' or (current_user.permissions and 'appointments' in current_user.permissions):
        today = get_now().date()
        today_appointments = Appointment.query.filter(Appointment.company_id == current_user.company_id, Appointment.date == today).all()
        if today_appointments:
            notifications.append({
                'type': 'info',
                'title': 'Citas Hoy',
                'message': f'Tienes {len(today_appointments)} citas programadas para hoy.',
                'link': url_for('appointments', date_filter='today')
            })
            
    return jsonify({
        'count': len(notifications),
        'notifications': notifications
    })

# ==========================================
# MÓDULO: CLIENTES (API)
# ==========================================
@app.route('/api/client/<document_id>')
@login_required
def get_client(document_id):
    client = Client.query.filter_by(document_id=document_id, company_id=current_user.company_id).first()
    if client:
        return jsonify({
            'success': True,
            'name': client.name,
            'phone': client.phone,
            'email': client.email,
            'address': client.address
        })
    return jsonify({'success': False})

# ==========================================
# MÓDULO: CITAS
# ==========================================
@app.route('/appointments')
@login_required
@permission_required('appointments')
def appointments():
    importance = request.args.get('importance')
    date_filter = request.args.get('date_filter')
    
    query = Appointment.query.filter_by(company_id=current_user.company_id)
    
    if importance:
        query = query.filter(Appointment.importance == importance)
    
    if date_filter == 'today':
        query = query.filter(Appointment.date == get_now().date())
    elif date_filter:
        try:
            target_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
            query = query.filter(Appointment.date == target_date)
        except ValueError:
            pass
            
    appointments = query.order_by(Appointment.date.asc(), Appointment.time.asc()).all()
    return render_template('appointments.html', appointments=appointments)

@app.route('/appointments/add', methods=['POST'])
@login_required
@permission_required('appointments')
def add_appointment():
    client_name = request.form.get('client_name')
    client_phone = request.form.get('client_phone')
    service = request.form.get('service')
    date_str = request.form.get('date')
    time_str = request.form.get('time')
    importance = request.form.get('importance', 'medium')

    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    time_obj = datetime.strptime(time_str, '%H:%M').time()

    new_appointment = Appointment(
        client_name=client_name,
        client_phone=client_phone,
        service=service,
        date=date_obj,
        time=time_obj,
        importance=importance,
        company_id=current_user.company_id
    )
    db.session.add(new_appointment)
    db.session.commit()
    flash('Cita agendada correctamente.', 'success')
    return redirect(url_for('appointments'))

@app.route('/appointments/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('appointments')
def delete_appointment(id):
    appointment = Appointment.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    db.session.delete(appointment)
    db.session.commit()
    flash('Cita eliminada correctamente.', 'success')
    return redirect(url_for('appointments'))


# ==========================================
# MÓDULO: FACTURACIÓN Y COTIZACIONES
# ==========================================
@app.route('/billing')
@login_required
@permission_required('billing')
def billing():
    open_register = CashRegister.query.filter_by(status='open', company_id=current_user.company_id, user_id=current_user.id).first()
    if not open_register:
        flash('Debes abrir tu caja antes de poder realizar ventas.', 'warning')
        return redirect(url_for('cash_register'))
    return render_template('billing.html')

@app.route('/billing/process', methods=['POST'])
@login_required
@permission_required('billing')
def process_billing():
    open_register = CashRegister.query.filter_by(status='open', company_id=current_user.company_id, user_id=current_user.id).first()
    if not open_register:
        return jsonify({'error': 'No tienes una caja abierta'}), 400

    data = request.get_json()
    items = data.get('items', [])
    tax_rate = float(data.get('tax_rate', 19.0))
    client_data = data.get('client', {})
    
    if not items:
        return jsonify({'error': 'No hay productos en la factura'}), 400
        
    doc_id = client_data.get('document_id')
    client_id = None
    if doc_id:
        client = Client.query.filter_by(document_id=doc_id, company_id=current_user.company_id).first()
        if not client:
            client = Client(
                document_id=doc_id,
                name=client_data.get('name', 'Consumidor Final'),
                phone=client_data.get('phone', ''),
                email=client_data.get('email', ''),
                address=client_data.get('address', ''),
                company_id=current_user.company_id
            )
            db.session.add(client)
            db.session.flush()
        client_id = client.id
        
    subtotal = 0.0
    new_invoice = Invoice(
        tax_rate=tax_rate,
        client_id=client_id,
        cash_register_id=open_register.id,
        company_id=current_user.company_id
    )
    db.session.add(new_invoice)
    db.session.flush()
    
    for item in items:
        product_id = item.get('product_id')
        qty = int(item.get('quantity', 1))
        
        product = Product.query.filter_by(id=product_id, company_id=current_user.company_id).first()
        if product:
            item_subtotal = product.sale_price * qty
            subtotal += item_subtotal
            
            invoice_item = InvoiceItem(
                invoice_id=new_invoice.id,
                product_id=product.id,
                quantity=qty,
                unit_price=product.sale_price,
                subtotal=item_subtotal
            )
            db.session.add(invoice_item)
            # Actualizar stock: Si tiene componentes (receta), descontar de los insumos
            if product.components:
                for comp in product.components:
                    comp_prod = Product.query.get(comp.component_product_id)
                    if comp_prod:
                        needed_qty = comp.quantity * qty
                        
                        # Lógica de Descuento por Lotes para Insumos
                        remaining_comp_qty = needed_qty
                        if comp_prod.track_stock:
                            comp_batches = ProductBatch.query.filter_by(product_id=comp_prod.id).order_by(ProductBatch.expiry_date.asc()).all()
                            for b in comp_batches:
                                if remaining_comp_qty <= 0: break
                                if b.quantity >= remaining_comp_qty:
                                    b.quantity -= remaining_comp_qty
                                    remaining_comp_qty = 0
                                else:
                                    remaining_comp_qty -= b.quantity
                                    b.quantity = 0
                            
                            comp_prod.stock -= needed_qty
                            db.session.add(InventoryMovement(
                                product_id=comp_prod.id,
                                movement_type='out',
                                quantity=needed_qty,
                                reason=f'Venta Factura #{new_invoice.id} (Insumo de {product.name})',
                                company_id=current_user.company_id,
                                user_id=current_user.id
                            ))
            else:
                # Lógica de Descuento por Lotes (FIFO)
                remaining_qty = qty
                if product.track_stock:
                    # Buscar lotes del producto ordenados por vencimiento
                    batches = ProductBatch.query.filter_by(product_id=product.id).order_by(ProductBatch.expiry_date.asc()).all()
                    for batch in batches:
                        if remaining_qty <= 0: break
                        if batch.quantity >= remaining_qty:
                            batch.quantity -= remaining_qty
                            remaining_qty = 0
                        else:
                            remaining_qty -= batch.quantity
                            batch.quantity = 0
                    
                    product.stock -= qty
                    db.session.add(InventoryMovement(
                        product_id=product.id,
                        movement_type='out',
                        quantity=qty,
                        reason=f'Venta Factura #{new_invoice.id}',
                        user_id=current_user.id,
                        company_id=current_user.company_id
                    ))
            
    tax_amount = subtotal * (tax_rate / 100.0)
    total = subtotal + tax_amount
    
    new_invoice.subtotal = subtotal
    new_invoice.tax_amount = tax_amount
    new_invoice.total = total
    
    db.session.commit()
    
    return jsonify({
        'success': True,
        'invoice_id': new_invoice.id,
        'message': 'Factura procesada correctamente'
    })

@app.route('/invoice/<int:id>/print')
@login_required
@permission_required('billing')
def print_invoice(id):
    invoice = Invoice.query.filter_by(id=id, company_id=current_user.company_id).first_or_404()
    return render_template('invoice_print.html', invoice=invoice)

@app.route('/clients')
@login_required
def clients_list():
    page = request.args.get('page', 1, type=int)
    pagination = Client.query.filter_by(company_id=current_user.company_id).order_by(Client.name.asc()).paginate(page=page, per_page=10, error_out=False)
    clients = pagination.items
    return render_template('clients.html', clients=clients, pagination=pagination)

@app.route('/sales')
@login_required
@permission_required('billing')
def sales_history():
    page = request.args.get('page', 1, type=int)
    # Obtener parámetros de filtro
    invoice_id = request.args.get('invoice_id')
    client_doc = request.args.get('client_doc')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    min_price = request.args.get('min_price')
    max_price = request.args.get('max_price')
    sort = request.args.get('sort', 'date_desc')

    # Query base
    query = Invoice.query.filter_by(company_id=current_user.company_id)

    # Aplicar filtros si existen
    if invoice_id:
        query = query.filter(Invoice.id == invoice_id)
    
    if client_doc:
        query = query.join(Client).filter(Client.document_id.like(f"%{client_doc}%"))
    
    if start_date:
        query = query.filter(Invoice.date >= start_date)
    
    if end_date:
        query = query.filter(Invoice.date <= end_date + " 23:59:59")
    
    if min_price:
        query = query.filter(Invoice.total >= float(min_price))
        
    if max_price:
        query = query.filter(Invoice.total <= float(max_price))

    # Aplicar ordenamiento
    if sort == 'date_asc':
        query = query.order_by(Invoice.date.asc())
    elif sort == 'total_asc':
        query = query.order_by(Invoice.total.asc())
    elif sort == 'total_desc':
        query = query.order_by(Invoice.total.desc())
    else: # date_desc por defecto
        query = query.order_by(Invoice.date.desc())

    invoices_pagination = query.paginate(page=page, per_page=10, error_out=False)
    invoices = invoices_pagination.items
    return render_template('sales_history.html', invoices=invoices, pagination=invoices_pagination)

# ==========================================
# MÓDULO: CAJA (CASH REGISTER)
# ==========================================
@app.route('/cash_register')
@login_required
@permission_required('cash')
def cash_register():
    # El usuario ve SU propia caja abierta
    open_register = CashRegister.query.filter_by(status='open', company_id=current_user.company_id, user_id=current_user.id).first()
    
    # El admin ve todo el historial, el empleado solo el suyo
    if current_user.role == 'admin':
        recent_registers = CashRegister.query.filter_by(company_id=current_user.company_id).order_by(CashRegister.id.desc()).limit(20).all()
        # Para el admin, mostramos quién abrió cada caja si es posible
    else:
        recent_registers = CashRegister.query.filter_by(company_id=current_user.company_id, user_id=current_user.id).order_by(CashRegister.id.desc()).limit(10).all()
    
    current_sales = 0
    if open_register:
        for inv in open_register.invoices:
            current_sales += inv.total
            
    return render_template('cash_register.html', 
                           open_register=open_register, 
                           recent_registers=recent_registers,
                           current_sales=current_sales)

@app.route('/cash_register/open', methods=['POST'])
@login_required
@permission_required('cash')
def open_register():
    if CashRegister.query.filter_by(status='open', company_id=current_user.company_id, user_id=current_user.id).first():
        flash('Ya tienes una caja abierta.', 'warning')
        return redirect(url_for('cash_register'))
    
    initial_balance = float(request.form.get('initial_balance', 0))
    new_register = CashRegister(
        initial_balance=initial_balance,
        company_id=current_user.company_id,
        user_id=current_user.id,
        status='open'
    )
    db.session.add(new_register)
    db.session.commit()
    flash('Caja abierta correctamente.', 'success')
    return redirect(url_for('cash_register'))

@app.route('/cash_register/close', methods=['POST'])
@login_required
@permission_required('cash')
def close_register():
    register = CashRegister.query.filter_by(status='open', company_id=current_user.company_id, user_id=current_user.id).first()
    if not register:
        flash('No tienes una caja abierta para cerrar.', 'danger')
        return redirect(url_for('cash_register'))
    
    final_balance = float(request.form.get('final_balance', 0))
    
    # Calcular ventas de esta caja específica
    current_sales = sum(inv.total for inv in register.invoices)
    
    register.final_balance = final_balance
    register.expected_balance = register.initial_balance + current_sales
    register.closed_at = get_now()
    register.status = 'closed'
    
    db.session.commit()
    flash('Caja cerrada correctamente.', 'success')
    return redirect(url_for('cash_register'))



# ==========================================
# MÓDULO: SUSCRIPCIÓN Y EXPORTACIÓN
# ==========================================
@app.route('/suspended')
@login_required
def suspended():
    company = current_user.company
    if not company or company.subscription_status != 'suspended':
        return redirect(url_for('dashboard'))
    return render_template('suspended.html', company=company)

@app.route('/export_data')
@login_required
def export_data():
    # Solo dueños/administradores pueden exportar todo
    if current_user.role != 'admin':
        flash('Solo el administrador de la empresa puede descargar los datos.', 'danger')
        return redirect(url_for('suspended'))
        
    memory_file = io.BytesIO()
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        # 1. Export Products
        products = Product.query.filter_by(company_id=current_user.company_id).all()
        prod_io = io.StringIO()
        writer = csv.writer(prod_io)
        writer.writerow(['ID', 'Nombre', 'Descripción', 'Precio Compra', 'Precio Venta', 'Stock'])
        for p in products:
            writer.writerow([p.id, p.name, p.description, p.purchase_price, p.sale_price, p.stock])
        zf.writestr('inventario.csv', prod_io.getvalue())
        
        # 2. Export Clients
        clients = Client.query.filter_by(company_id=current_user.company_id).all()
        client_io = io.StringIO()
        writer = csv.writer(client_io)
        writer.writerow(['ID', 'Documento', 'Nombre', 'Teléfono', 'Email', 'Dirección'])
        for c in clients:
            writer.writerow([c.id, c.document_id, c.name, c.phone, c.email, c.address])
        zf.writestr('clientes.csv', client_io.getvalue())
        
        # 3. Export Sales (Invoices)
        invoices = Invoice.query.filter_by(company_id=current_user.company_id).all()
        inv_io = io.StringIO()
        writer = csv.writer(inv_io)
        writer.writerow(['Factura ID', 'Fecha', 'Cliente ID', 'Subtotal', 'Impuestos', 'Total'])
        for inv in invoices:
            writer.writerow([inv.id, inv.date, inv.client_id, inv.subtotal, inv.tax_amount, inv.total])
        zf.writestr('ventas.csv', inv_io.getvalue())
        
    memory_file.seek(0)
    return send_file(memory_file, download_name='mis_datos_erp.zip', as_attachment=True)

@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    if current_user.role != 'admin':
        flash('Solo el administrador puede acceder a la configuración de la empresa.', 'danger')
        return redirect(url_for('dashboard'))
    
    company = current_user.company
    if request.method == 'POST':
        company.name = request.form.get('name')
        company.document_id = request.form.get('document_id')
        company.phone = request.form.get('phone')
        company.email = request.form.get('email')
        company.address = request.form.get('address')
        company.primary_color = request.form.get('primary_color')
        company.secondary_color = request.form.get('secondary_color')
        
        # Logo upload handling
        file = request.files.get('logo')
        if file and file.filename != '':
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            company.logo_path = 'uploads/' + filename
            
        db.session.commit()
        flash('Configuración de la empresa actualizada correctamente.', 'success')
        return redirect(url_for('settings'))
        
    return render_template('settings.html', company=company)

@app.route('/god-mode', methods=['GET', 'POST'])
def system_admin():
    companies = Company.query.all()
    if request.method == 'POST':
        password = request.form.get('password')
        if password == 'SoyElCreador123': # Contraseña maestra
            action = request.form.get('action')
            
            if action == 'create_company':
                email = request.form.get('email')
                user_password = request.form.get('user_password')
                if User.query.filter_by(email=email).first():
                    flash('El correo ya está registrado.', 'danger')
                else:
                    new_company = Company(name='Empresa Nueva (Por configurar)')
                    db.session.add(new_company)
                    db.session.flush()
                    
                    new_admin = User(
                        email=email,
                        name='Administrador ' + email.split('@')[0],
                        password_hash=generate_password_hash(user_password),
                        role='admin',
                        permissions='all',
                        company_id=new_company.id,
                        must_change_password=True
                    )
                    db.session.add(new_admin)
                    db.session.commit()
                    flash(f'Empresa y usuario creados para {email}', 'success')
                    return redirect(url_for('system_admin'))

            company_id = request.form.get('company_id')
            company = Company.query.get(company_id)
            if not company:
                flash('Empresa no seleccionada.', 'warning')
                return redirect(url_for('system_admin'))

            if action == 'renew':
                try:
                    months = int(request.form.get('months', 1))
                except ValueError:
                    months = 1
                    
                import calendar
                def add_months(sourcedate, m):
                    month = sourcedate.month - 1 + m
                    year = sourcedate.year + month // 12
                    month = month % 12 + 1
                    day = min(sourcedate.day, calendar.monthrange(year, month)[1])
                    return sourcedate.replace(year=year, month=month, day=day)
                
                if company.subscription_status == 'active' and company.valid_until > get_now():
                    new_date = add_months(company.valid_until, months)
                else:
                    new_date = add_months(get_now(), months)
                    
                company.valid_until = new_date
                company.subscription_status = 'active'
                db.session.commit()
                flash(f'Suscripción de {company.name} extendida hasta el {new_date.strftime("%d-%m-%Y")}.', 'success')
            elif action == 'suspend':
                company.subscription_status = 'suspended'
                company.valid_until = get_now() - timedelta(days=1)
                db.session.commit()
                flash(f'Empresa {company.name} suspendida.', 'warning')
        else:
            flash('Contraseña incorrecta.', 'danger')
            
    return render_template('system_admin.html', companies=companies)

@app.route('/setup', methods=['GET', 'POST'])
@login_required
def setup_company():
    company = current_user.company
    if company.document_id: # Ya configurada
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        company.name = request.form.get('name')
        company.document_id = request.form.get('document_id')
        company.phone = request.form.get('phone')
        company.email = request.form.get('email')
        company.address = request.form.get('address')
        db.session.commit()
        flash('¡Configuración inicial completada! Bienvenido.', 'success')
        return redirect(url_for('dashboard'))
        
    return render_template('setup_wizard.html')

@app.route('/legal')
def legal_info():
    return render_template('legal.html')

@app.route('/export/inventory')
@login_required
def export_inventory():
    products = Product.query.filter_by(company_id=current_user.company_id).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Headers
    headers = ['ID', 'Producto', 'Descripción', 'Precio Compra', 'Precio Venta', 'Stock']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    for p in products:
        row = [p.id, p.name, p.description, p.purchase_price, p.sale_price, p.stock]
        ws.append(row)
        # Format prices
        ws.cell(row=ws.max_row, column=4).number_format = '"$"#,##0'
        ws.cell(row=ws.max_row, column=5).number_format = '"$"#,##0'
        
    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, 
                     download_name=f"inventario_{get_now().strftime('%Y%m%d')}.xlsx")

@app.route('/export/sales')
@login_required
def export_sales():
    invoice_id = request.args.get('invoice_id')
    client_doc = request.args.get('client_doc')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    min_price = request.args.get('min_price')
    max_price = request.args.get('max_price')
    sort = request.args.get('sort', 'date_desc')
    
    query = Invoice.query.filter_by(company_id=current_user.company_id)
    
    if invoice_id:
        query = query.filter(Invoice.id == invoice_id)
    if client_doc:
        query = query.join(Client).filter(Client.document_id.like(f"%{client_doc}%"))
    if start_date:
        query = query.filter(Invoice.date >= start_date)
    if end_date:
        query = query.filter(Invoice.date <= end_date + " 23:59:59")
    if min_price:
        query = query.filter(Invoice.total >= float(min_price))
    if max_price:
        query = query.filter(Invoice.total <= float(max_price))

    # Aplicar ordenamiento
    if sort == 'date_asc':
        query = query.order_by(Invoice.date.asc())
    elif sort == 'total_asc':
        query = query.order_by(Invoice.total.asc())
    elif sort == 'total_desc':
        query = query.order_by(Invoice.total.desc())
    else: # date_desc por defecto
        query = query.order_by(Invoice.date.desc())
        
    invoices = query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    
    # Headers
    headers = ['Factura ID', 'Fecha', 'Cliente', 'Subtotal', 'Impuestos', 'Total']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    for i in invoices:
        client_name = i.client.name if i.client else 'Consumidor Final'
        row = [i.id, i.date.strftime('%Y-%m-%d %H:%M'), client_name, i.subtotal, i.tax_amount, i.total]
        ws.append(row)
        # Format totals
        ws.cell(row=ws.max_row, column=4).number_format = '"$"#,##0'
        ws.cell(row=ws.max_row, column=5).number_format = '"$"#,##0'
        ws.cell(row=ws.max_row, column=6).number_format = '"$"#,##0'
        
    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, 
                     download_name=f"ventas_{get_now().strftime('%Y%m%d')}.xlsx")

if __name__ == '__main__':
    app.run(debug=True)
